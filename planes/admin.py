from django.contrib import admin
from django.http import HttpResponse
from django.utils.html import format_html
from django.forms.models import BaseInlineFormSet
from collections import defaultdict
from django.urls import path, reverse
from django.shortcuts import render
from django.db.models import Sum, Count

from .models import Plan, ResumenPlan, Potrero, PracticaPotrero, EvaluacionTecnica
from .pdf_utils import generar_pdf_constancia

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import openpyxl
import csv
from pyproj import Transformer


# ======================
# EXPORTAR EXCEL
# ======================
def exportar_excel(modeladmin, request, queryset):
    from openpyxl import Workbook
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Planes"

    headers = [
        'Ranking',
        'Numero',
        'RUT',
        'Agricultor',
        'Operador',
        'Tipo concurso',

        'Comuna',
        'Sector',

        'Rol avaluo',
        'Tenencia',
        'Superficie predio',
        'Superficie intervenida',

        'Coord Norte',
        'Coord Este',
        'Huso',

        'Potreros',

        'Costo practicas',
        'Costo analisis',
        'Costo asesoria',
        'Costo total',

        'Incentivo practicas',
        'Total incentivo',

        'Puntaje tecnico',
        'Estado final',

        'Estado administrativo',
        'Motivo rechazo admin',

        'Estado tecnico',
        'Motivo rechazo tecnico',

        'Estado reconsideracion',
        'Motivo reconsideracion',
        'Participación',
    ]

    ws.append(headers)

    # ======================
    # FUNCIÓN ESTADO FINAL
    # ======================
    def obtener_estado_final(resumen):
        if not resumen:
            return ''

        # Reconsideración manda
        if resumen.estado_reconsideracion == 'rechazado':
            return 'rechazado'

        if resumen.estado_reconsideracion == 'aprobado':
            return 'aprobado'

        # Administrativo
        if resumen.estado_administrativo == 'rechazado':
            return 'rechazado'

        # Técnico
        if resumen.estado_tecnico == 'rechazado':
            return 'rechazado'

        # Todo aprobado
        if (
            resumen.estado_administrativo == 'aprobado' and
            resumen.estado_tecnico == 'aprobado'
        ):
            return 'aprobado'

        return 'pendiente'

    # ======================
    # DATOS
    # ======================
    ranking = list(modeladmin.get_queryset(request))

    for index, plan in enumerate(ranking, start=1):
        resumen = ResumenPlan.objects.filter(plan=plan).first()

        if not resumen:
            print(f"⚠️ Plan {plan.numero} sin resumen")

        print(f"Plan {plan.numero} - resumen:", resumen)

        from planes.models import HistorialPostulacion

        rut = str(plan.rut_agricultor).replace(".", "").strip()
        historial = HistorialPostulacion.objects.filter(rut__iexact=rut).first()

        veces = historial.veces if historial else 0

        if veces == 0:
            participacion_txt = "Primera vez"
        elif veces == 1:
            participacion_txt = "Segunda vez"
        else:
            participacion_txt = "Tercera vez"

        ws.append([
            index,
            plan.numero,
            plan.rut_agricultor,
            plan.nombre_agricultor,
            plan.nombre_operador,
            plan.concurso,

            plan.comuna,
            plan.sector,

            getattr(resumen, 'rol_avaluo', '') if resumen else '',
            getattr(resumen, 'tenencia', '') if resumen else '',
            getattr(resumen, 'superficie_total', '') if resumen else '',
            getattr(resumen, 'superficie_potreros', '') if resumen else '',

            getattr(resumen, 'coordenada_norte', '') if resumen else '',
            getattr(resumen, 'coordenada_este', '') if resumen else '',
            getattr(resumen, 'huso', '') if resumen else '',

            ", ".join([p.nombre for p in plan.potreros.all()]),

            getattr(resumen, 'costo_practicas', 0) if resumen else 0,
            getattr(resumen, 'costo_analisis', 0) if resumen else 0,
            getattr(resumen, 'costo_asesoria', 0) if resumen else 0,
            getattr(resumen, 'costo_total_real', 0) if resumen else 0,

            getattr(resumen, 'incentivo_practicas', 0) if resumen else 0,
            getattr(resumen, 'incentivo_total', 0) if resumen else 0,

            getattr(resumen, 'puntaje_tecnico', 0) if resumen else 0,
            obtener_estado_final(resumen),

            getattr(resumen, 'estado_administrativo', '') if resumen else '',
            getattr(resumen, 'motivo_rechazo_admin', '') if resumen else '',

            getattr(resumen, 'estado_tecnico', '') if resumen else '',
            getattr(resumen, 'motivo_rechazo_tecnico', '') if resumen else '',

            getattr(resumen, 'estado_reconsideracion', '') if resumen else '',
            getattr(resumen, 'motivo_reconsideracion', '') if resumen else '',
            participacion_txt,
        ])


    # ======================
    # HOJA 2 — DETALLE POTREROS
    # ======================
    ws2 = wb.create_sheet(title="Detalle Potreros")

    headers2 = [
        'Numero plan',
        'RUT',
        'Agricultor',
        'Operador',
        'Tipo concurso',

        'Potrero',
        'Superficie',

        'Practicas',

        'Coord Norte',
        'Coord Este',
        'Huso',
    ]

    ws2.append(headers2)

    for plan in ranking:
        resumen = ResumenPlan.objects.filter(plan=plan).first()

        for potrero in plan.potreros.all():

            practicas = ", ".join([
                f"{p.tipo}-{p.subtipo_enmienda}" if p.subtipo_enmienda else p.tipo
                for p in potrero.practicas.all()
            ])

            ws2.append([
                plan.numero,
                plan.rut_agricultor,
                plan.nombre_agricultor,
                plan.nombre_operador,
                plan.concurso,

                potrero.nombre,
                potrero.superficie,

                practicas,

                resumen.coordenada_norte if resumen else '',
                resumen.coordenada_este if resumen else '',
                resumen.huso if resumen else '',
            ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=planes_completo.xlsx'

    wb.save(response)
    return response


exportar_excel.short_description = "Exportar a Excel"
# ======================
# EXPORTAR PUNTAJES
# ======================
def exportar_puntajes(modeladmin, request, queryset):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Puntajes"

    headers = [
        "Plan", "RUT", "Nombre", "Concurso",
        "Fósforo (ppm)", "Potasio (cmol/kg)", "Azufre", "Saturación Al (%)", "Superficie praderas",
        "Puntaje Fósforo", "Puntaje Potasio", "Puntaje Azufre",
        "Puntaje Cal", "Puntaje Praderas", "Puntaje Aporte", "Puntaje Participación",
        "TOTAL",
    ]

    ws.append(headers)

    for plan in queryset:

        ev = getattr(plan, 'evaluaciontecnica', None)
        if not ev:
            continue

        fosforo = potasio = azufre = aluminio = None
        superficie = 0

        for potrero in plan.potreros.all():
            superficie += potrero.superficie or 0

            for practica in potrero.practicas.all():

                if practica.tipo == 'fosforo' and fosforo is None:
                    fosforo = practica.nivel_inicial

                if practica.tipo == 'enmienda':
                    if practica.subtipo_enmienda == 'potasio' and potasio is None:
                        potasio = practica.nivel_inicial

                    if practica.subtipo_enmienda == 'azufre' and azufre is None:
                        azufre = practica.nivel_inicial

                    if practica.subtipo_enmienda == 'cal' and aluminio is None:
                        aluminio = practica.saturacion_aluminio

        ws.append([
            plan.numero,
            plan.rut_agricultor,
            plan.nombre_agricultor,
            plan.concurso,
            fosforo, potasio, azufre, aluminio, superficie,
            ev.puntaje_fosforo(),
            ev.puntaje_potasio(),
            ev.puntaje_azufre(),
            ev.puntaje_cal(),
            ev.puntaje_pradera(),
            ev.puntaje_aporte_financiero(),
            ev.puntaje_participacion(),
            ev.puntaje,
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=puntajes.xlsx'

    wb.save(response)
    return response

exportar_puntajes.short_description = "Exportar puntajes"


# ======================
# DASHBOARD
# ======================
def dashboard_view(request):

    planes = Plan.objects.all().distinct()

    from planes.models import ResumenPlan
    from collections import Counter

    resumenes = ResumenPlan.objects.all()

    total = sum([r.incentivo_total or 0 for r in resumenes])

    def obtener_estado_final(resumen):

        if not resumen:
            return ''

        if resumen.estado_reconsideracion == 'rechazado':
            return 'rechazado'

        if resumen.estado_reconsideracion == 'aprobado':
            return 'aprobado'

        if resumen.estado_administrativo == 'rechazado':
            return 'rechazado'

        if resumen.estado_tecnico == 'RECHAZADO':
            return 'rechazado'

        if (
            resumen.estado_administrativo == 'aprobado' and
            resumen.estado_tecnico == 'APROBADO'
        ):
            return 'aprobado'

        return 'pendiente'

    aprobados = sum([
        r.incentivo_total or 0
        for r in resumenes
        if obtener_estado_final(r) == 'aprobado'
    ])

    rechazados_admin = sum([
        r.incentivo_total or 0
        for r in resumenes
        if obtener_estado_final(r) == 'rechazado'
    ])

    rechazados_tec = 0
    

    # AGRUPACIONES
    por_comuna = Plan.objects.exclude(comuna__isnull=True).exclude(comuna='').values('comuna').annotate(
        total=Count('numero', distinct=True)
    ).order_by('-total')

    por_operador = Plan.objects.exclude(nombre_operador__isnull=True).exclude(nombre_operador='').values('nombre_operador').annotate(
        total=Count('numero', distinct=True)
    ).order_by('-total')

    # CONCURSOS (el que usa tu template)
    concursos = [
        (p.concurso or '').strip().lower()
        for p in planes
    ]

    conteo_concurso = Counter([
        c for c in concursos if c
    ])
    print("DEBUG CONCURSO:", conteo_concurso)

    montos = {
        'total': total,
        'aprobados': aprobados,
        'rechazados_tecnicos': rechazados_tec,
        'rechazados_admin': rechazados_admin,
    }

    context = {
        'por_comuna': {
            item['comuna']: item['total']
            for item in por_comuna
        },
        'por_operador': {
            item['nombre_operador']: item['total']
            for item in por_operador
        },
        'conteo_concurso': dict(conteo_concurso),
        'planes_por_practica': {
            item['tipo']: item['total']
            for item in PracticaPotrero.objects.values('tipo').annotate(total=Count('id'))
        },
        'montos': {
            'total': total,
            'aprobados': aprobados,
            'rechazados_tecnicos': rechazados_tec,
            'rechazados_admin': rechazados_admin,
        }
    }
    return render(request, 'admin/dashboard.html', context)


# ======================
# INLINE RESUMEN
# ======================
class ResumenPlanInline(admin.StackedInline):
    model = ResumenPlan
    extra = 0
    max_num = 1
    can_delete = False

    def has_add_permission(self, request, obj=None):
        if obj and ResumenPlan.objects.filter(plan=obj).exists():
            return False
        return True
    
    def ver_mapa(self, obj):
        if obj and obj.coordenada_norte is not None and obj.coordenada_este is not None and obj.huso:

            try:
                norte = float(str(obj.coordenada_norte).replace(",", "."))
                este = float(str(obj.coordenada_este).replace(",", "."))

                transformer = Transformer.from_crs(
                    f"EPSG:327{int(obj.huso)}",
                    "EPSG:4326",
                    always_xy=True
                )

                lon, lat = transformer.transform(este, norte)

                url = f"https://www.google.com/maps?q={lat},{lon}"

                return format_html(
                    '<a target="_blank" href="{}">📍 Ver en Maps</a>', url
                )

            except Exception:
                return "Error en coordenadas"

        return "Guardar para ver mapa"

    ver_mapa.short_description = "Mapa"
    readonly_fields = ('ver_mapa',)
    fields = (
            'correo',
            'telefono',

            'rol_avaluo',
            'tenencia',

            'superficie_total',

            'coordenada_norte',
            'coordenada_este',
            'huso',
            'ver_mapa',  
    )        
# ======================
# INLINE POTRERO
# ======================
class PotreroInline(admin.StackedInline):
    model = Potrero
    extra = 0
    readonly_fields = ('ir_a_potrero',)

    def ir_a_potrero(self, obj):
        if obj.pk:
            return format_html(
                '<a href="/admin/{}/{}/{}/change/">Abrir potrero</a>',
                obj._meta.app_label,
                obj._meta.model_name,
                obj.pk
            )
        return "-"
    ir_a_potrero.short_description = "Editar prácticas"

    def ver_mapa(self, obj):
        if obj and obj.utm_norte and obj.utm_este and obj.huso:

            try:
                norte = float(str(obj.utm_norte).replace(",", "."))
                este = float(str(obj.utm_este).replace(",", "."))

                transformer = Transformer.from_crs(
                    f"EPSG:327{int(obj.huso)}",
                    "EPSG:4326",
                    always_xy=True
                )

                lon, lat = transformer.transform(este, norte)

                url = f"https://www.google.com/maps?q={lat},{lon}"

                return format_html(
                    '<a target="_blank" href="{}">📍 Ver en Maps</a>', url
                )

            except Exception as e:
                return f"Error: {e}"

        return "Guardar para ver mapa"

    ver_mapa.short_description = "Mapa"
    readonly_fields = ('ir_a_potrero', 'ver_mapa')
    fields = (
        'nombre',
        'superficie',

        'utm_norte',
        'utm_este',
        'huso',
        'ver_mapa',

        'ir_a_potrero',
)

# ======================
# FORM PRACTICA DINAMICO
# ======================
from django import forms

class PracticaPotreroForm(forms.ModelForm):

    class Meta:
        model = PracticaPotrero
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # ocultar todo primero
        self.fields['subtipo_enmienda'].widget = forms.Select(
            choices=self.fields['subtipo_enmienda'].choices
        )
        self.fields['subtipo_cubierta'].widget = forms.Select(
            choices=self.fields['subtipo_cubierta'].choices
        )
        self.fields['saturacion_aluminio'].widget = forms.NumberInput()
        # 🔥 IMPORTANTE: asegurar que nivel_inicial SIEMPRE exista
        self.fields['nivel_inicial'].widget = forms.NumberInput()
        
        tipo = self.data.get('tipo') or getattr(self.instance, 'tipo', None)

        if tipo and (
                'enmienda' in str(tipo).lower() or
                'incorporacion' in str(tipo).lower()
            ):
                choices = [
                    c for c in self.fields['subtipo_enmienda'].choices
                    if c[0] != ''
                ]

                self.fields['subtipo_enmienda'].widget = forms.Select(
                    choices=choices
                )

                self.fields['saturacion_aluminio'].widget = forms.NumberInput()

                subtipo = None

                if self.instance and self.instance.pk:
                    subtipo = self.instance.subtipo_enmienda
                else:
                    subtipo = self.data.get('subtipo_enmienda')

                # 🔥 SOLO CAL oculta nivel inicial
                if subtipo == 'cal':
                    self.fields['nivel_inicial'].required = False
                    

                else:
                    self.fields['nivel_inicial'].required = True
                    self.fields['nivel_inicial'].widget = forms.NumberInput()

        

        # lógica por tipo
        if tipo and (
            'enmienda' in str(tipo).lower() or
            'incorporacion' in str(tipo).lower()
        ):
            choices = [
                c for c in self.fields['subtipo_enmienda'].choices
                if c[0] != ''
            ]

            self.fields['subtipo_enmienda'].widget = forms.Select(
                choices=choices
            )

            self.fields['saturacion_aluminio'].widget = forms.NumberInput()

            self.fields['nivel_inicial'].required = False
            
            
        elif tipo == 'cubierta':
            self.fields['subtipo_cubierta'].widget = forms.Select(
                choices=self.fields['subtipo_cubierta'].choices
            )
            self.fields['subtipo_cubierta'].required = True

        elif tipo == 'fosforo':
            pass  # solo niveles

# ======================
# INLINE PRACTICAS
# ======================
class PracticaPotreroInline(admin.StackedInline):
    model = PracticaPotrero
    form = PracticaPotreroForm
    extra = 0

    class Media:
        js = ('planes/js/practicas.js',)

    def get_formset(self, request, obj=None, **kwargs):
        formset = super().get_formset(request, obj, **kwargs)

        # 🔥 esto muestra errores reales
        formset.form.base_fields['nivel_inicial'].required = False

        return formset

# ======================
# INLINE EVALUACION
# ======================
class EvaluacionTecnicaInline(admin.StackedInline):
    model = EvaluacionTecnica
    extra = 1
    readonly_fields = ('puntaje', 'ver_desglose')

    def ver_desglose(self, obj):
        if not obj:
            return "-"
        return f"Total: {obj.puntaje}"

    ver_desglose.short_description = "Detalle"


@admin.register(Plan)
class PlanAdmin(admin.ModelAdmin):

    class Media:
        css = {
            'all': ('admin_custom.css',)
        }

    inlines = [ResumenPlanInline, PotreroInline, EvaluacionTecnicaInline]

    

    readonly_fields = ('boton_constancia',)

    # 🔎 COLUMNAS
    list_display = (
        'ranking_posicion',
        'numero',
        'nombre_agricultor',
        'rut_agricultor',
        'comuna',
        'nombre_operador',
        'estado_admin_formateado',
        'estado_tecnico_formateado',
        'puntaje_tecnico',
        'incentivo_total_display',
        'boton_constancia',
    )

    def incentivo_total(self, obj):
        from .models import ResumenPlan

        resumen = ResumenPlan.objects.filter(plan=obj).first()

        if resumen and resumen.incentivo_total:
            return int(resumen.incentivo_total)

        return "-"
    def incentivo_total_display(self, obj):
        from .models import ResumenPlan

        resumen = ResumenPlan.objects.filter(plan=obj).first()

        if resumen and resumen.incentivo_total:
            return int(resumen.incentivo_total)

        return "-"
    incentivo_total_display.short_description = "INCENTIVO $"
    # 🔍 BUSCADOR (NUEVO)
    search_fields = (
        'numero',
        'nombre_agricultor',
        'rut_agricultor',
        'comuna',
        'nombre_operador',
    )

    def participacion_display(self, obj):
        return obj.get_participacion_agricultor_display()

    participacion_display.short_description = "Participación"

    # 🧰 FILTROS (NUEVO)
    list_filter = (
        'comuna',
        'concurso',
        'nombre_operador',
        'estado_administrativo',
        'evaluaciontecnica__estado_tecnico',
    )

    # ↕️ ORDEN BASE (NUEVO)
    ordering = ('-evaluaciontecnica__puntaje',)

    # 📄 PAGINACIÓN (NUEVO)
    list_per_page = 20

    # ⚡ PERFORMANCE (NUEVO)
    list_select_related = ('evaluaciontecnica',)

    def estado_tecnico_formateado(self, obj):
        ev = getattr(obj, 'evaluaciontecnica', None)
        if not ev or not ev.estado_tecnico:
            return "-"
        return ev.estado_tecnico.upper()

    estado_tecnico_formateado.short_description = "Estado técnico"


    def estado_admin_formateado(self, obj):
        if not obj.estado_administrativo:
            return "-"
        return obj.estado_administrativo.upper()

    estado_admin_formateado.short_description = "Estado admin"

    # ======================
    # ACCIONES
    # ======================
    actions = [
        exportar_excel,
        exportar_puntajes,
        'exportar_ranking_csv'
    ]

    # ======================
    # CAMPOS CALCULADOS
    # ======================
    def estado_admin_display(self, obj):
        return obj.estado_administrativo

    def estado_tecnico(self, obj):
        if hasattr(obj, 'evaluaciontecnica'):
            return obj.evaluaciontecnica.estado_tecnico
        return "-"

    def puntaje_tecnico(self, obj):
        if hasattr(obj, 'evaluaciontecnica'):
            return obj.evaluaciontecnica.puntaje
        return "-"

    # ======================
    # QUERYSET (RANKING)
    # ======================
    def get_ranking_queryset(self, request):
        return Plan.objects.filter(
            evaluaciontecnica__estado_tecnico='aprobado'
        ).order_by('-evaluaciontecnica__puntaje')

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.order_by('-fecha_ingreso', '-evaluaciontecnica__puntaje')
    
    # ======================
    # RANKING
    # ======================
    def ranking_posicion(self, obj):
        try:
            cl = self.get_changelist_instance(self.request)
            queryset = list(cl.result_list)

            for index, plan in enumerate(queryset):
                if plan.pk == obj.pk:
                    return index + 1

        except Exception as e:
            return "-"

        return "-"
    
    def changelist_view(self, request, extra_context=None):
        self.request = request
        return super().changelist_view(request, extra_context)
    
    ordering = ('-evaluaciontecnica__puntaje', '-fecha_ingreso')
    # ======================
    # EXPORTAR CSV
    # ======================
    def exportar_ranking_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response.write('\ufeff')
        response['Content-Disposition'] = 'attachment; filename="ranking_planes.csv"'

        writer = csv.writer(response, delimiter=';')

        writer.writerow(['Ranking', 'Numero', 'Agricultor', 'Puntaje'])

        ranking = list(self.get_queryset(request))

        for index, plan in enumerate(ranking, start=1):
            writer.writerow([
                index,
                plan.numero,
                plan.nombre_agricultor,
                plan.evaluaciontecnica.puntaje if plan.evaluaciontecnica else 0
            ])

        return response

    exportar_ranking_csv.short_description = "Exportar ranking CSV"

    def boton_constancia(self, obj):
        if obj.pk:
            url = reverse('ver_constancia_pdf', args=[obj.pk])
            return format_html(
                '<a class="boton-pdf" target="_blank" href="{}">📄 Constancia</a>',
                url
            )
        return "-"

    boton_constancia.short_description = "Constancia"

    def incentivo_total(self, obj):
        resumen = getattr(obj, 'resumenplan', None)
        if resumen:
            return resumen.incentivo_total
        return "-"
    
    incentivo_total.short_description = "Incentivo $"

    
    def save_model(self, request, obj, form, change):
        if not obj.usuario:
            obj.usuario = request.user
        super().save_model(request, obj, form, change)

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)

        plan = form.instance

        from .models import ResumenPlan

        resumen, _ = ResumenPlan.objects.get_or_create(plan=plan)
        resumen.save()

# ======================
# ADMIN RESUMEN
# ======================
@admin.register(ResumenPlan)
class ResumenPlanAdmin(admin.ModelAdmin):
    exclude = ('tipo_postulacion',)
    readonly_fields = (
        'ver_mapa',
        'nombres_potreros',
        'superficie_potreros',
        'costo_practicas',
        'costo_analisis',
        'costo_asesoria',
        'costo_total_real',
        'incentivo_practicas',
        'incentivo_total',
        'estado_tecnico',
        'puntaje_tecnico',
        'estado_administrativo',
        'motivo_rechazo_admin',
        'motivo_rechazo_tecnico',
        'estado_reconsideracion',
        'motivo_reconsideracion',
    )
    fields = (
        'plan',

        'correo',
        'telefono',

        'rol_avaluo',
        'tenencia',

        'superficie_total',
        'superficie_potreros',

        'coordenada_este',
        'coordenada_norte',
        'huso',
        'ver_mapa', 

        'nombres_potreros',

        'costo_practicas',
        'costo_analisis',
        'costo_asesoria',
        'costo_total_real',

        'incentivo_practicas',
        'incentivo_total',

        'estado_administrativo',
        'motivo_rechazo_admin',

        'estado_tecnico',
        'motivo_rechazo_tecnico',  

        'estado_reconsideracion',
        'motivo_reconsideracion',

        'puntaje_tecnico', 
    )
    from pyproj import Transformer

    from pyproj import Transformer

    def ver_mapa(self, obj):
        if obj and obj.coordenada_norte and obj.coordenada_este and obj.huso:

            try:
                # 🔥 corregir formato (coma → punto)
                norte = float(str(obj.coordenada_norte).replace(",", "."))
                este = float(str(obj.coordenada_este).replace(",", "."))

                transformer = Transformer.from_crs(
                    f"EPSG:327{int(obj.huso)}",
                    "EPSG:4326",
                    always_xy=True
                )

                lon, lat = transformer.transform(este, norte)

                url = f"https://www.google.com/maps?q={lat},{lon}"

                return format_html(
                    '<a target="_blank" href="{}">📍 Ver en Maps</a>', url
                )

            except Exception as e:
                return f"Error: {e}"

        return "Sin coordenadas"

    ver_mapa.short_description = "Mapa"
   

    list_display = ('plan', 'nombres_potreros')

# ======================
# ADMIN POTRERO
# ======================
@admin.register(Potrero)
class PotreroAdmin(admin.ModelAdmin):
    inlines = [PracticaPotreroInline]


admin.site.site_header = "Programa Apoyo al Mejoramiento de la Fertilidad en Sistemas Agropecuarios Productivos Región de Los Ríos"
admin.site.site_title = "SEREMI de Agricultura - Región de Los Ríos"
admin.site.index_title = "Tercer Concurso • 2026"
admin.site.site_url = "/admin/dashboard/"

from django.contrib.admin import AdminSite
from django.db.models import Sum, Count
from .models import Plan, PracticaPotrero


class MyAdminSite(AdminSite):
    site_header = "Administración Concurso"

    def index(self, request, extra_context=None):
        extra_context = extra_context or {}

        # ======================
        # BASE
        # ======================
        planes = Plan.objects.all()
        practicas = PracticaPotrero.objects.all()

        # ======================
        # KPIs (MONTOS)
        # ======================
        extra_context['total_ingresado'] = practicas.aggregate(
            total=Sum('costo')
        )['total'] or 0

        extra_context['aprobados_monto'] = practicas.filter(
            potrero__plan__evaluaciontecnica__estado_tecnico='aprobado'
        ).aggregate(total=Sum('costo'))['total'] or 0

        extra_context['rechazados_tecnicos_monto'] = practicas.filter(
            potrero__plan__evaluaciontecnica__estado_tecnico='rechazado'
        ).aggregate(total=Sum('costo'))['total'] or 0

        extra_context['rechazados_admin_monto'] = practicas.filter(
            potrero__plan__estado_administrativo='rechazado'
        ).aggregate(total=Sum('costo'))['total'] or 0

        # ======================
        # KPIs (CANTIDAD)
        # ======================
        extra_context['total_planes'] = planes.count()

        extra_context['aprobados'] = planes.filter(
            estado_administrativo='aprobado'
        ).count()

        extra_context['rechazados_admin'] = planes.filter(
            estado_administrativo='rechazado'
        ).count()

        extra_context['rechazados_tecnicos'] = planes.filter(
            evaluaciontecnica__estado_tecnico='rechazado'
        ).count()

        # ======================
        # AGRUPACIONES
        # ======================

        por_comuna = planes.values('comuna').annotate(
            total=Count('numero', distinct=True)
        ).order_by('-total')

        extra_context['planes_por_comuna'] = {
            item['comuna']: item['total']
            for item in por_comuna
        }

        extra_context['planes_por_operador'] = {
            item['nombre_operador']: item['total']
            for item in planes.values('nombre_operador').annotate(total=Count('numero', distinct=True))
        }

        extra_context['planes_por_concurso'] = {
            item['concurso']: item['total']
            for item in planes.values('concurso').annotate(total=Count('numero', distinct=True))
        }

        extra_context['planes_por_practica'] = {
            item['tipo']: item['total']
            for item in practicas.values('tipo').annotate(total=Count('numero', distinct=True))
        }

        return super().index(request, extra_context)

# ======================
# DASHBOARD URL (SEGURO)
# ======================
from django.urls import path

def get_admin_urls():
    urls = original_get_urls()

    custom_urls = [
        path(
            'dashboard/',
            admin.site.admin_view(dashboard_view),
            name='dashboard'
        ),
    ]

    return custom_urls + urls


original_get_urls = admin.site.get_urls
admin.site.get_urls = get_admin_urls


from .models import Agenda

@admin.register(Agenda)
class AgendaAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'correo', 'fecha', 'hora', 'creado')
